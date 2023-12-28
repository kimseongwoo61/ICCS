# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime
import os

# 동작 로깅용도
import logging
logger = logging.getLogger('django')

# 엑셀에 유해성 검사 결과를 엑셀 파일에 저장한다.
# input : title, url, domain, check_blackList, image_dirInfo, summary, score_illegal, username
# return : 유해성 검사 정보가 저장된 엑셀파일
def save_data(output_dir, title, url, domain, check_blackList, summary, image_dirInfo, score_illegal, username):
    logger.info("[*] save_data - START")
    
    # 엑셀 파일 로드
    workbook = load_workbook(output_dir)
    
    # 시트 이름 지정
    sheet_name = "Sheet1"
    
    # 특정 시트 가져오기 (이어 작성할 시트)
    sheet = workbook[sheet_name]
    
    
    # 우선 보고서 작성자의 기본 프로필 정보를 삽입한다.
    reporter = sheet['B3'].value
    write_time = sheet['D3'].value
    
    if (reporter is None) or (write_time is None):
        current_time = datetime.datetime.now()
        time_string = current_time.strftime("%Y-%m-%d %H:%M:%S")

        sheet['B3'] = username
        sheet['D3'] = time_string
    
    
    # 콘텐츠 유해성 검사 결과를 엑셀 파일에 기록하도록 한다.
    # 처음 시작 셀 번호는 5번임.
    start_row = find_last_row(output_dir, "Sheet1", 1) + 1
    
    # 작성 작업 수행
    sheet['A' + str(start_row)] = title
    sheet['B' + str(start_row)] = url
    sheet['C' + str(start_row)] = domain
    sheet['D' + str(start_row)] = check_blackList
    sheet['G' + str(start_row)] = summary
    
    likelihood_name = (
        "UNKNOWN",
        "VERY_UNLIKELY",
        "UNLIKELY",
        "POSSIBLE",
        "LIKELY",
        "VERY_LIKELY",
    )
    

    if(len(image_dirInfo) > 0):
        for i in range(0, len(image_dirInfo)):
            score =  f"adult: {likelihood_name[score_illegal[i].adult]}  "
            score += f"medical: {likelihood_name[score_illegal[i].medical]}  "
            score += f"spoofed: {likelihood_name[score_illegal[i].spoof]}  "
            score += f"violence: {likelihood_name[score_illegal[i].violence]}  "
            score += f"racy: {likelihood_name[score_illegal[i].racy]}"
            

            # 정규식 패턴과 매치되는 부분을 추출
            file_name_with_extension = os.path.basename(image_dirInfo[i])

            if file_name_with_extension:
                sheet['E' + str(start_row + i)].font = Font(bold=True,color='FF0000')
                sheet['E' + str(start_row + i)] = """=HYPERLINK("./image/""" + file_name_with_extension + """", "유해 이미지_주의")"""
                
            else:
                # 확장자를 정상적으로 추출할 수 없는 경우
                logger.warning("[?] save_data - ERROR : 확장자를 정상적으로 추출할 수 없습니다.")
                sheet['E' + str(start_row + i)] = "None"
            
            
            
            sheet['F' + str(start_row + i)] = score

    # 변경 사항 저장
    workbook.save(output_dir)
    logger.info("[*] save_data - FINISH")
    
    
# 저장된 엑셀 파일 중에서 가장 마지막 행 정보를 찾아 반환한다.
# input : 파일 경로(절대경로), 시트 이름, 시작 행
# return : 컬럼중에서 가장 마지막에 존재하는 행 번호를 반환한다.
def find_last_row(filename, sheet_name, start_row):
    logger.info("[*] find_last_row - START")
    # 엑셀 파일 로드
    workbook = load_workbook(filename)

    # 특정 시트 가져오기
    sheet = workbook[sheet_name]

    # A~G 열 순회
    max_row = start_row - 1  # 가장 마지막 행 번호를 저장할 변수
    for column in range(ord('A'), ord('H')):
        column_letter = chr(column)
        
        # 열의 데이터를 거꾸로 순회하며 가장 마지막 행 번호를 찾음
        for row in range(sheet.max_row, start_row - 1, -1):
            cell = sheet[column_letter + str(row)]
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                break
    
    logger.info("[*] find_last_row - FINISH")
    return max_row

    